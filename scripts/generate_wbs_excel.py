#!/usr/bin/env python3
"""
WBS → Excel & Gantt 차트 생성 스크립트
wbs.json 데이터를 읽어:
1. WBS 상세 테이블 (시트1) - 각 작업의 상세 정보
2. 업무별 요약 (시트2) - 단계/주별 그룹 요약
3. 간트차트 (시트3) - 타임라인 시각화

산출물:
    LifeLog_WBS.xlsx — 3개 시트 (WBS 상세 / 업무별 요약 / 간트차트)
"""

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

WBS_FILE = Path(__file__).resolve().parent.parent / "wbs.json"
OUTPUT_FILE = Path(__file__).resolve().parent.parent / "LifeLog_WBS.xlsx"

def load_wbs():
    if not WBS_FILE.exists():
        print(f"ERROR: WBS file not found: {WBS_FILE}")
        sys.exit(1)
    with open(WBS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def parse_date(d):
    if not d:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(d.strip(), fmt)
        except ValueError:
            continue
    return None

def style_header_row(ws, row_num, cols):
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def apply_cell_style(cell, bold=False, fill=None, align="left", wrap=True, font_size=10, border=True):
    cell.font = Font(name="Arial", size=font_size, bold=bold)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    return cell

def priority_color(p):
    return {
        "critical": "FF6B6B",
        "high": "FFB347",
        "normal": "7CB9E8",
        "low": "A9DFBF"
    }.get(p, "D3D3D3")

def create_wbs_sheet(ws, wbs_data):
    headers = [
        "WBS ID", "Level", "Parent", "Task Name", "Description",
        "Milestone", "Start Date", "End Date", "Duration (주)",
        "Assignees", "Priority", "Deliverable", "Labels"
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))

    wbs_levels = wbs_data.get("wbs_levels", {})
    rows = []
    for wbs_id, item in wbs_levels.items():
        if wbs_id == "WP0":
            continue
        start = parse_date(item.get("start_date", ""))
        end = parse_date(item.get("end_date", ""))
        weeks = item.get("estimated_weeks", 0)
        rows.append({
            "id": wbs_id,
            "level": item.get("level", 0),
            "parent": item.get("parent_id", ""),
            "name": item.get("name", ""),
            "desc": item.get("desc", ""),
            "milestone": item.get("milestone", ""),
            "start": start,
            "end": end,
            "weeks": weeks,
            "assignees": item.get("assignees", []),
            "priority": item.get("priority", "normal"),
            "deliverable": item.get("deliverable", ""),
            "labels": item.get("labels", [])
        })

    phase_fills = {
        "M-01": "BDD7EE",
        "M-02": "D5A6BD",
        "M-03": "FFF2CC",
        "M-04": "B4C6E7",
        "M-05": "E2EFDA",
        "M-06": "FCE4D6"
    }

    for i, row_data in enumerate(rows):
        row_num = i + 2
        phase_fill = PatternFill(start_color=phase_fills.get(row_data["milestone"], "FFFFFF"),
                                 end_color=phase_fills.get(row_data["milestone"], "FFFFFF"), fill_type="solid")
        priority_fill = PatternFill(start_color=priority_color(row_data["priority"]),
                                    end_color=priority_color(row_data["priority"]), fill_type="solid")
        
        ws.cell(row=row_num, column=1, value=row_data["id"]).font = Font(bold=True)
        apply_cell_style(ws.cell(row=row_num, column=1), fill=phase_fill)
        apply_cell_style(ws.cell(row=row_num, column=2), fill=phase_fill)
        ws.cell(row=row_num, column=2, value=row_data["level"])
        apply_cell_style(ws.cell(row=row_num, column=3), fill=phase_fill)
        ws.cell(row=row_num, column=3, value=row_data["parent"])
        apply_cell_style(ws.cell(row=row_num, column=4, value=row_data["name"]), fill=phase_fill)
        apply_cell_style(ws.cell(row=row_num, column=5, value=row_data["desc"]), fill=phase_fill)
        apply_cell_style(ws.cell(row=row_num, column=6, value=row_data["milestone"]), fill=phase_fill)
        
        if row_data["start"]:
            apply_cell_style(ws.cell(row=row_num, column=7, value=row_data["start"].strftime("%Y-%m-%d")), fill=phase_fill)
        else:
            apply_cell_style(ws.cell(row=row_num, column=7, value=""), fill=phase_fill)
        
        if row_data["end"]:
            apply_cell_style(ws.cell(row=row_num, column=8, value=row_data["end"].strftime("%Y-%m-%d")), fill=phase_fill)
        else:
            apply_cell_style(ws.cell(row=row_num, column=8, value=""), fill=phase_fill)
        
        apply_cell_style(ws.cell(row=row_num, column=9, value=int(row_data["weeks"])), fill=phase_fill)
        apply_cell_style(ws.cell(row=row_num, column=10, value=", ".join(row_data["assignees"])), fill=phase_fill)
        priority_cell = apply_cell_style(ws.cell(row=row_num, column=11, value=row_data["priority"]), fill=priority_fill)
        priority_cell.font = Font(bold=True, color="000000" if row_data["priority"] != "critical" else "FFFFFF")
        apply_cell_style(ws.cell(row=row_num, column=12, value=row_data["deliverable"]), fill=phase_fill)
        apply_cell_style(ws.cell(row=row_num, column=13, value=", ".join(row_data["labels"])), fill=phase_fill)

    col_widths = [12, 8, 12, 45, 60, 12, 12, 12, 10, 20, 12, 30, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_summary_sheet(ws, wbs_data):
    milestones = wbs_data.get("milestones", [])
    wbs_levels = wbs_data.get("wbs_levels", {})

    headers = ["단계 ID", "단계명", "시작일", "종료일", "기간 (주)",
               "Tasks", "Critical Tasks", "Assignees", "Deliverables"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))

    for i, m in enumerate(milestones):
        row_num = i + 2
        start = parse_date(m.get("start_date", ""))
        end = parse_date(m.get("due_date", ""))
        weeks = round((end - start).days / 7) if start and end else 0

        task_items = [v for k, v in wbs_levels.items() if v.get("milestone") == m["id"] or v.get("milestone_override") == m["id"]]
        children_keys = set()
        for k, t in wbs_levels.items():
            if t.get("milestone") == m["id"] or t.get("milestone_override") == m["id"]:
                if len(t.get("children", [])) > 0:
                    children_keys.update(t["children"])
                else:
                    children_keys.add(k)
        
        deliverables = [t.get("deliverable", "") for t in task_items if t.get("deliverable")]
        for ck in children_keys:
            cd = wbs_levels.get(ck, {})
            if cd.get("deliverable"):
                deliverables.append(cd["deliverable"])
        
        assignees = set()
        for t in task_items + [wbs_levels.get(c) for c in children_keys if wbs_levels.get(c)]:
            if isinstance(t, dict):
                assignees.update(t.get("assignees", []))

        ws.cell(row=row_num, column=1, value=m["id"]).font = Font(bold=True)
        phase_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") 
        for col_idx in range(1, len(headers) + 1):
            apply_cell_style(ws.cell(row=row_num, column=col_idx), fill=phase_fill)
        
        ws.cell(row=row_num, column=1, value=m["id"])
        ws.cell(row=row_num, column=2, value=m["gitlab_title"])
        if start:
            ws.cell(row=row_num, column=3, value=start.strftime("%Y-%m-%d"))
        if end:
            ws.cell(row=row_num, column=4, value=end.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=5, value=weeks)
        ws.cell(row=row_num, column=6, value=len(children_keys))
        ws.cell(row=row_num, column=7, value=len([t for t in task_items if t.get("priority") == "critical"]))
        ws.cell(row=row_num, column=8, value=", ".join(sorted(assignees)))
        ws.cell(row=row_num, column=9, value="\n".join([d for d in deliverables if d]))

    col_widths = [12, 40, 12, 12, 10, 8, 15, 30, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_gantt_sheet(ws, wbs_data):
    milestones = wbs_data.get("milestones", [])
    wbs_levels = wbs_data.get("wbs_levels", {})
    
    if not milestones:
        print("ERROR: milestones not found in wbs.json")
        sys.exit(1)

    start_date = parse_date(min(m["start_date"] for m in milestones))
    end_date = parse_date(max(m["due_date"] for m in milestones))
    
    weeks = []
    current = start_date
    while current <= end_date:
        weeks.append(current)
        current += timedelta(weeks=1)
    
    total_weeks = len(weeks)
    print(f"Generating Gantt: {total_weeks} weeks from {start_date.strftime('%Y-%m-%d')} to {(end_date - timedelta(days=1)).strftime('%Y-%m-%d')}")
    
    gantt_header_cols = ["Task ID", "Task Name"]
    gantt_data_cols = total_weeks
    
    for col_idx, header in enumerate(gantt_header_cols, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, 2 + total_weeks)
    
    phase_fills = {
        "M-01": "BDD7EE",
        "M-02": "D5A6BD",
        "M-03": "FFF2CC",
        "M-04": "B4C6E7",
        "M-05": "E2EFDA",
        "M-06": "FCE4D6"
    }
    
    for i, m in enumerate(milestones):
        task_items = [v for k, v in wbs_levels.items() if v.get("milestone") == m["id"] or v.get("milestone_override") == m["id"]]
        rows = []
        for t in task_items:
            if t.get("name") and t.get("start_date") and t.get("end_date"):
                t_start = parse_date(t["start_date"])
                t_end = parse_date(t["end_date"])
                if t_start and t_end and min(weeks) <= t_start <= max(weeks):
                    rows.append((t.get("id", ""), t.get("name", ""), t.get("milestone", ""), t_start, t_end))
        ws.cell(row=i+2, column=1, value=m["id"])
        ws.cell(row=i+2, column=2, value=f"■ {m['gitlab_title']}")
        for col_idx in range(1, 2 + total_weeks):
            apply_cell_style(ws.cell(row=i+2, column=col_idx), font_size=10, bold=True)
        
        for j, week_date in enumerate(weeks):
            bar = False
            ws.cell(row=i+2, column=2+j)
            if min(weeks) <= week_date <= max(weeks):
                bar = True
            if bar:
                fill = PatternFill(start_color=phase_fills.get(m["id"], "4472C4"),
                                  end_color=phase_fills.get(m["id"], "4472C4"), fill_type="solid")
                for col_idx in range(1, 2 + total_weeks):
                    cell = ws.cell(row=i+2, column=col_idx)
                    if col_idx > 2:
                        cell.fill = fill
    
    current_row = len(milestones) + 3
    
    for i, m in enumerate(milestones):
        phase_fill = PatternFill(start_color=phase_fills.get(m["id"], "4472C4"),
                                 end_color=phase_fills.get(m["id"], "4472C4"), fill_type="solid")
        task_items = [v for k, v in wbs_levels.items() if v.get("milestone") == m["id"] or v.get("milestone_override") == m["id"]]
        children_keys = set()
        for k, t in wbs_levels.items():
            if t.get("milestone") == m["id"] or t.get("milestone_override") == m["id"]:
                if len(t.get("children", [])) > 0:
                    children_keys.update(t["children"])
                else:
                    children_keys.add(k)
        
        for child_key in children_keys:
            child = wbs_levels.get(child_key)
            if not child or not child.get("start_date") or not child.get("end_date"):
                continue
            child_start = parse_date(child["start_date"])
            child_end = parse_date(child["end_date"])
            if not child_start or not child_end:
                continue
            
            ws.cell(row=current_row, column=1, value=child["id"])
            ws.cell(row=current_row, column=2, value=f"  ├ {child.get('name', child['id'])}")
            for col_idx in range(1, 2 + total_weeks):
                apply_cell_style(ws.cell(row=current_row, column=col_idx), fill=phase_fill)
            
            for j, week_date in enumerate(weeks):
                bar_start = child_start
                bar_end = child_end
                if (min(weeks) <= week_date <= max(weeks)):
                    if bar_start <= week_date <= bar_end + timedelta(weeks=1):
                        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        ws.cell(row=current_row, column=2+j).fill = fill
            
            current_row += 1

    print(f"Generated Gantt chart for {current_row - (len(milestones) + 3)} tasks")

def main():
    print("Loading WBS data...")
    wbs_data = load_wbs()
    print(f"Loaded {len(wbs_data.get('milestones', []))} milestones, {len(wbs_data.get('wbs_levels', {}))} WBS items")
    print(f"\nGenerating Excel file: {OUTPUT_FILE}")
    
    wb = Workbook()
    
    print("Creating WBS detailed table...")
    ws_wbs = wb.active
    ws_wbs.title = "WBS 상세"
    create_wbs_sheet(ws_wbs, wbs_data)
    
    print("Creating summary table...")
    ws_summary = wb.create_sheet("업무별 요약")
    create_summary_sheet(ws_summary, wbs_data)
    
    print("Creating Gantt chart...")
    ws_gantt = wb.create_sheet("간트차트")
    create_gantt_sheet(ws_gantt, wbs_data)
    
    wb.save(OUTPUT_FILE)
    print(f"\n✓ Excel file saved to: {OUTPUT_FILE}")
    print(f"  - WBS 상세 (시트1)")
    print(f"  - 업무별 요약 (시트2)")
    print(f"  - 간트차트 (시트3)")

if __name__ == "__main__":
    main()